"""Reads an Excel workbook and groups rows into ZipGroup objects."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import structlog

from app.models import Role, StructuralAnomaly, ZipGroup

log = structlog.get_logger()

_PLACEHOLDER = re.compile(r"^\[to be filled\]$", re.IGNORECASE)


@dataclass
class _ColumnMap:
    zip_code: int
    gov_level: int
    branch: int
    role: int
    mycivx_name: int
    mycivx_website: int


def _header_words(header: str) -> str:
    """Strip non-alphanumeric chars and return lowercase word string."""
    cleaned = re.sub(r"[^\w\s]", " ", header, flags=re.UNICODE)
    return re.sub(r"\s+", " ", cleaned).lower().strip()


def _find_col(headers: dict[int, str], *keywords: str) -> int | None:
    """Return the column index whose header contains all given keywords."""
    for col_idx, raw in headers.items():
        words = _header_words(raw)
        if all(kw in words for kw in keywords):
            return col_idx
    return None


def _detect_columns(sheet: openpyxl.worksheet.worksheet.Worksheet) -> _ColumnMap:
    headers: dict[int, str] = {}
    for col in sheet.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value is not None:
                headers[cell.column] = str(cell.value)

    def require(col: int | None, label: str) -> int:
        if col is None:
            raise ValueError(f"Required column not found in workbook: {label!r}")
        return col

    return _ColumnMap(
        zip_code=require(_find_col(headers, "zip"), "Zip Code"),
        gov_level=require(_find_col(headers, "gov"), "Gov Level"),
        branch=require(_find_col(headers, "branch"), "Branch"),
        role=require(_find_col(headers, "role"), "Role"),
        mycivx_name=require(_find_col(headers, "mycivx", "name"), "MyCivX – Name"),
        mycivx_website=require(_find_col(headers, "mycivx", "website"), "MyCivX – Website"),
    )


def _cell_value(row: tuple, col: int) -> str | None:
    """Return a non-empty, non-placeholder cell value or None."""
    raw = row[col - 1].value
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or _PLACEHOLDER.match(s):
        return None
    return s


def read_zip_groups(workbook_path: Path) -> tuple[list[ZipGroup], list[StructuralAnomaly]]:
    """Read an Excel workbook and return ZIP groups and any structural anomalies.

    ZIP code cells are forward-filled in memory — the workbook is never modified.

    Args:
        workbook_path: Path to the .xlsx input file.

    Returns:
        A tuple of (zip_groups, anomalies).
    """
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = wb.active
    cols = _detect_columns(sheet)

    groups: dict[str, tuple[list[Role], list[int]]] = {}
    order: list[str] = []
    current_zip: str | None = None
    anomalies: list[StructuralAnomaly] = []

    for row in sheet.iter_rows(min_row=2):
        zip_val = _cell_value(row, cols.zip_code)
        if zip_val:
            current_zip = zip_val

        if current_zip is None:
            log.warning("excel_reader.row_before_first_zip", row=row[0].row)
            continue

        role_val = _cell_value(row, cols.role)
        if not role_val:
            anomalies.append(StructuralAnomaly(
                zip_code=current_zip,
                actual_role_count=len(groups.get(current_zip, ([], []))[0]),
                notes=f"Row {row[0].row} has no role value",
            ))
            log.warning("excel_reader.row_without_role", zip=current_zip, row=row[0].row)
            continue

        role = Role(
            name=role_val,
            gov_level=_cell_value(row, cols.gov_level) or "",
            branch=_cell_value(row, cols.branch) or "",
            mycivx_name=_cell_value(row, cols.mycivx_name),
            mycivx_website=_cell_value(row, cols.mycivx_website),
        )

        if current_zip not in groups:
            groups[current_zip] = ([], [])
            order.append(current_zip)

        groups[current_zip][0].append(role)
        groups[current_zip][1].append(row[0].row)

    result = [
        ZipGroup(zip_code=z, roles=roles, row_indices=indices)
        for z in order
        for roles, indices in (groups[z],)
    ]

    log.info("excel_reader.complete", zip_count=len(result), anomaly_count=len(anomalies))
    return result, anomalies
