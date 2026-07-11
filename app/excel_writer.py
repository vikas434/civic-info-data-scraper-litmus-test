"""Writes LLM results into an Excel workbook, preserving all formatting."""

from __future__ import annotations

import re
from pathlib import Path
from types import TracebackType

import openpyxl
import structlog

from app.config import ModelConfig
from app.matcher import names_match, urls_match
from app.models import MatchResult, ZipGroup, ZipResponse

log = structlog.get_logger()


def _header_words(header: str) -> str:
    cleaned = re.sub(r"[^\w\s]", " ", header, flags=re.UNICODE)
    return re.sub(r"\s+", " ", cleaned).lower().strip()


def _find_model_columns(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    column_prefix: str,
) -> tuple[int, int, int]:
    """Locate the Name, Website, and Match? columns for a given model prefix.

    Args:
        sheet: The active worksheet.
        column_prefix: The prefix used in column headers, e.g. "Gemini 2.5 Pro".

    Returns:
        A tuple of (name_col, website_col, match_col) as 1-based column indices.

    Raises:
        ValueError: If any of the three columns cannot be found.
    """
    prefix_words = _header_words(column_prefix)
    name_col = website_col = match_col = None

    for col in sheet.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value is None:
                continue
            words = _header_words(str(cell.value))
            if prefix_words not in words:
                continue
            if "name" in words:
                name_col = cell.column
            elif "website" in words or "url" in words:
                website_col = cell.column
            elif "match" in words:
                match_col = cell.column

    missing = []
    if name_col is None:
        missing.append(f"{column_prefix} – Name")
    if website_col is None:
        missing.append(f"{column_prefix} – Website")
    if match_col is None:
        missing.append(f"{column_prefix} – Match?")

    if missing:
        raise ValueError(f"Missing columns in workbook: {missing}")

    return name_col, website_col, match_col  # type: ignore[return-value]


class WorkbookWriter:
    """Context manager that holds a workbook open for multiple ZIP writes.

    Usage::

        with WorkbookWriter(path, model) as writer:
            for zip_group, response in results:
                match_results = writer.write_zip(zip_group, response)
    """

    def __init__(self, path: Path, model: ModelConfig) -> None:
        self._path = path
        self._model = model
        self._wb: openpyxl.Workbook | None = None
        self._name_col: int = 0
        self._website_col: int = 0
        self._match_col: int = 0

    def __enter__(self) -> WorkbookWriter:
        self._wb = openpyxl.load_workbook(self._path)
        sheet = self._wb.active
        self._name_col, self._website_col, self._match_col = _find_model_columns(
            sheet, self._model.column_prefix
        )
        log.info(
            "workbook_writer.opened",
            path=str(self._path),
            model=self._model.id,
            name_col=self._name_col,
            website_col=self._website_col,
            match_col=self._match_col,
        )
        return self

    def __exit__(
        self,
        exc_type: type[BaseException] | None,
        exc_val: BaseException | None,
        exc_tb: TracebackType | None,
    ) -> None:
        if self._wb is not None:
            self._wb.save(self._path)
            log.info("workbook_writer.saved", path=str(self._path))

    def write_zip(self, zip_group: ZipGroup, response: ZipResponse) -> list[MatchResult]:
        """Write LLM results for one ZIP group into the workbook.

        Args:
            zip_group: The ZIP group with roles and row indices.
            response: The structured LLM response for this ZIP.

        Returns:
            A list of MatchResult comparing LLM output to MyCivX ground truth.
        """
        if self._wb is None:
            raise RuntimeError("write_zip called outside context manager")
        sheet = self._wb.active
        results_by_role = {r.role: r for r in response.results}
        match_results: list[MatchResult] = []

        for role, row_idx in zip(zip_group.roles, zip_group.row_indices, strict=True):
            result = results_by_role.get(role.name)
            if result is None:
                log.warning(
                    "workbook_writer.missing_role",
                    zip=zip_group.zip_code,
                    role=role.name,
                )
                continue

            sheet.cell(row=row_idx, column=self._name_col).value = result.official_name
            sheet.cell(row=row_idx, column=self._website_col).value = result.official_website

            has_ground_truth = bool(role.mycivx_name and role.mycivx_website)
            n_match = (
                names_match(role.mycivx_name, result.official_name) if role.mycivx_name else False
            )
            w_match = (
                urls_match(role.mycivx_website, result.official_website)
                if role.mycivx_website
                else False
            )
            cell_match = has_ground_truth and n_match and w_match

            sheet.cell(row=row_idx, column=self._match_col).value = cell_match

            match_results.append(MatchResult(
                role=role.name,
                expected_name=role.mycivx_name or "",
                expected_website=role.mycivx_website or "",
                actual_name=result.official_name,
                actual_website=result.official_website,
                name_match=n_match,
                website_match=w_match,
            ))

        log.info(
            "workbook_writer.zip_written",
            zip=zip_group.zip_code,
            roles_written=len(match_results),
        )
        return match_results
