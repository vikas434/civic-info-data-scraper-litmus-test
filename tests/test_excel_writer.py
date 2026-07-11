"""Tests for Excel workbook writer."""

from pathlib import Path

import openpyxl
import pytest

from app.config import ModelConfig, ModelPricing
from app.excel_writer import WorkbookWriter
from app.models import Role, RoleResult, ZipGroup, ZipResponse


@pytest.fixture()
def model() -> ModelConfig:
    return ModelConfig(
        id="google/gemini-2.5-pro",
        display_name="Gemini 2.5 Pro",
        column_prefix="Gemini 2.5 Pro",
        has_web_search=True,
        pricing=ModelPricing(input_per_1k_tokens=0.00125, output_per_1k_tokens=0.010),
    )


def make_workbook_with_model_cols(tmp_path: Path, column_prefix: str) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "Zip Code", "Gov Level", "Branch", "Role",
        "MyCivX \u2013 Name", "MyCivX \u2013 Website",
        f"{column_prefix} \u2013 Name",
        f"{column_prefix} \u2013 Website",
        f"{column_prefix} \u2013 Match?",
    ])
    ws.append(["10019", "State", "Executive", "Governor",
               "Kathy Hochul", "https://www.ny.gov/governor", None, None, None])
    ws.append([None, "Local", "Executive", "Mayor",
               "Eric Adams", "https://www.nyc.gov", None, None, None])
    path = tmp_path / "workbook.xlsx"
    wb.save(path)
    return path


@pytest.fixture()
def workbook(tmp_path: Path, model: ModelConfig) -> Path:
    return make_workbook_with_model_cols(tmp_path, model.column_prefix)


def make_zip_group() -> ZipGroup:
    return ZipGroup(
        zip_code="10019",
        roles=[
            Role(name="Governor", gov_level="State", branch="Executive",
                 mycivx_name="Kathy Hochul", mycivx_website="https://www.ny.gov/governor"),
            Role(name="Mayor", gov_level="Local", branch="Executive",
                 mycivx_name="Eric Adams", mycivx_website="https://www.nyc.gov"),
        ],
        row_indices=[2, 3],
    )


def make_zip_response(governor_name: str = "Kathy Hochul") -> ZipResponse:
    return ZipResponse(
        zip_code="10019",
        results=[
            RoleResult(role="Governor", official_name=governor_name,
                       official_website="https://www.ny.gov/governor",
                       confidence=0.95, sources=["https://www.ny.gov"]),
            RoleResult(role="Mayor", official_name="Eric Adams",
                       official_website="https://www.nyc.gov",
                       confidence=0.90, sources=["https://www.nyc.gov"]),
        ],
    )


def read_cell(path: Path, row: int, col: int) -> object:
    wb = openpyxl.load_workbook(path)
    return wb.active.cell(row=row, column=col).value


def test_writes_name_to_correct_cell(workbook: Path, model: ModelConfig) -> None:
    with WorkbookWriter(workbook, model) as writer:
        writer.write_zip(make_zip_group(), make_zip_response())
    assert read_cell(workbook, 2, 7) == "Kathy Hochul"


def test_writes_website_to_correct_cell(workbook: Path, model: ModelConfig) -> None:
    with WorkbookWriter(workbook, model) as writer:
        writer.write_zip(make_zip_group(), make_zip_response())
    assert read_cell(workbook, 2, 8) == "https://www.ny.gov/governor"


def test_writes_match_true_when_correct(workbook: Path, model: ModelConfig) -> None:
    with WorkbookWriter(workbook, model) as writer:
        writer.write_zip(make_zip_group(), make_zip_response())
    assert read_cell(workbook, 2, 9) is True


def test_writes_match_false_when_wrong_name(workbook: Path, model: ModelConfig) -> None:
    with WorkbookWriter(workbook, model) as writer:
        writer.write_zip(make_zip_group(), make_zip_response(governor_name="Wrong Person"))
    assert read_cell(workbook, 2, 9) is False


def test_writes_second_role_row(workbook: Path, model: ModelConfig) -> None:
    with WorkbookWriter(workbook, model) as writer:
        writer.write_zip(make_zip_group(), make_zip_response())
    assert read_cell(workbook, 3, 7) == "Eric Adams"


def test_returns_match_results(workbook: Path, model: ModelConfig) -> None:
    with WorkbookWriter(workbook, model) as writer:
        match_results = writer.write_zip(make_zip_group(), make_zip_response())
    assert len(match_results) == 2
    assert match_results[0].role == "Governor"
    assert match_results[0].is_match is True


def test_does_not_touch_mycivx_columns(workbook: Path, model: ModelConfig) -> None:
    with WorkbookWriter(workbook, model) as writer:
        writer.write_zip(make_zip_group(), make_zip_response())
    # MyCivX Name col=5, Website col=6 must remain unchanged
    assert read_cell(workbook, 2, 5) == "Kathy Hochul"
    assert read_cell(workbook, 2, 6) == "https://www.ny.gov/governor"


def test_missing_role_in_response_is_skipped(workbook: Path, model: ModelConfig) -> None:
    response = ZipResponse(
        zip_code="10019",
        results=[
            RoleResult(role="Governor", official_name="Kathy Hochul",
                       official_website="https://www.ny.gov/governor",
                       confidence=0.95, sources=[]),
            # Mayor is missing
        ],
    )
    with WorkbookWriter(workbook, model) as writer:
        match_results = writer.write_zip(make_zip_group(), response)
    # Only Governor was written
    assert len(match_results) == 1
    assert read_cell(workbook, 3, 7) is None


def test_no_ground_truth_yields_no_match(workbook: Path, model: ModelConfig) -> None:
    group = ZipGroup(
        zip_code="10019",
        roles=[
            Role(name="Governor", gov_level="State", branch="Executive",
                 mycivx_name=None, mycivx_website=None),
        ],
        row_indices=[2],
    )
    response = ZipResponse(
        zip_code="10019",
        results=[
            RoleResult(role="Governor", official_name="Kathy Hochul",
                       official_website="https://www.ny.gov/governor",
                       confidence=0.90, sources=[]),
        ],
    )
    with WorkbookWriter(workbook, model) as writer:
        match_results = writer.write_zip(group, response)
    assert match_results[0].is_match is False
